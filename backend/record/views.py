from flask import jsonify, request, json
from func.records import Inbound, Outbound
from func.goods import Goods
from record import record_blue
import logging
from datetime import datetime
import os
from flask import send_file
from openpyxl import load_workbook

# logs
logging.basicConfig(filename="api.log", level=logging.DEBUG)


# Route to fetch the inbound records corresponding to a user ID
@record_blue.route('/get_inbound_record_by_user_id/<string:userID>')
def get_inbound_record_by_user_id(userID):
    try:
        result = Inbound.get_inbound_record_by_user_id(userID)
        logging.info('获得' + userID + '的入库操作记录信息')
        return jsonify(result)
    except Exception as e:
        logging.error(
            'An error occurred while retrieving inbound records by user ID from the database. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to fetch the outbound records corresponding to a user ID
@record_blue.route('/get_outbound_record_by_user_id/<string:userID>')
def get_outbound_record_by_user_id(userID):
    try:
        result = Outbound.get_outbound_record_by_user_id(userID)
        logging.info('获得' + userID + '的出库操作记录信息')
        return jsonify(result)
    except Exception as e:
        logging.error(
            'An error occurred while retrieving outbound records by user ID from the database. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to fetch all inbound records
@record_blue.route('/get_all_inbounds')
def get_all_inbounds():
    try:
        result = Inbound.get_all_inbounds()
        logging.info('获得所有入库操作记录信息')
        return jsonify(result)
    except Exception as e:
        logging.error(
            'An error occurred while retrieving all inbound records from the database. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to fetch all outbound records
@record_blue.route('/get_all_outbounds')
def outbounds():
    try:
        result = Outbound.get_all_outbounds()
        logging.info('获得所有出库操作记录信息')
        return jsonify(result)
    except Exception as e:
        logging.error(
            'An error occurred while retrieving all outbound records from the database. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to fetch the maximum outbound ID (used to generate a new outbound ID when creating a new outbound)
@record_blue.route('/get_max_outboundOrderID')
def get_max_outboundOrderID():
    try:
        result = Outbound.get_max_outboundOrderID()
        logging.info("获得最大的出库单号")
        return jsonify(result)
    except Exception as e:
        logging.error(
            "An error occurred while retrieving the maximum outbound order ID from the database. Error message: {}".format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to fetch the maximum outbound record ID (used to generate a new outbound record ID when creating a new outbound)
@record_blue.route('/get_max_outboundID')
def get_max_outboundID():
    try:
        result = Outbound.get_max_outboundID()
        logging.info("获得最大的出库记录ID")
        return jsonify(result)
    except Exception as e:
        logging.error(
            "An error occurred while retrieving the maximum outbound ID from the database. Error message: {}".format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to add a new outbound record
@record_blue.route('/add_outbound', methods=['GET', 'POST'])
def add_outbound():
    try:
        if request.method == 'POST':
            data = json.loads(request.get_data())
            result = Outbound.add_outbound(
                data['outboundID'],
                data['outboundOrderID'],
                data['outboundGoodsID'],
                data['outboundAmount'],
                data['outboundUpdatedByID'],
                datetime.now(),
            )
            return jsonify(result)
        else:
            return jsonify({'status': 'GET'})
    except Exception as e:
        logging.error(
            'An error occurred while adding the outbound record. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to generate and automatically download the outbound corresponding to the outbound order ID
@record_blue.route('/generate_outbound_order_by_id/<string:outboundOrderID>')
def generate_outbound_order_by_id(outboundOrderID):
    try:
        filename = Outbound.generate_outbound_order_by_id(outboundOrderID)
        if filename:
            # Get the current working directory
            current_dir = os.getcwd()

            parent_dir = os.path.dirname(current_dir)
            order_dir = os.path.join(parent_dir, 'order')
            # Create the "order" directory if it doesn't exist
            if not os.path.exists(order_dir):
                os.makedirs(order_dir)
            filepath = os.path.join(order_dir, filename)

            return send_file(filepath, as_attachment=True)
        else:
            return jsonify(
                {"error": "No records found for the given outbound order ID"}
            )
    except Exception as e:
        logging.error(
            'An error occurred while generating the outbound order by outbound order ID from the database. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route for confirming inbound information
@record_blue.route('/inboundConfirm', methods=['GET', 'POST'])
def inboundConfirm():
    try:
        if request.method == 'POST':
            goods_list = []
            userID = request.form.get('userID')
            file = request.files['file']

            # Get the current working directory
            current_dir = os.getcwd()
            parent_dir = os.path.dirname(current_dir)
            order_dir = os.path.join(parent_dir, 'order')
            # Create the "order" directory if it doesn't exist
            if not os.path.exists(order_dir):
                os.makedirs(order_dir)
            filepath = os.path.join(order_dir, file.filename)
            file.save(filepath)

            # Extraction of entry order data
            wb = load_workbook(filepath)
            ws = wb.active
            # Set the first row as the header row
            headers = [cell.value for cell in ws[1]]
            # Read data from the file line by line
            for row in ws.iter_rows(min_row=2, values_only=True):
                inbound_data = dict(zip(headers, row))
                goods_list.append(
                    {
                        'goodsID': '',
                        'goodsName': inbound_data['货物名称'],
                        'goodsSpecification': inbound_data['货物规格'],
                        'goodsManufacturer': inbound_data['生产厂商'],
                        'goodsProductionLicense': inbound_data['生产许可证'],
                        'goodsUnit': inbound_data['计量单位'],
                        'goodsAmount': inbound_data['入库数量'],
                        'goodsStorageCondition': inbound_data['存储条件'],
                        'goodsUpdatedByID': userID,
                        'goodsUpdatedTime': '',
                    }
                )
            return jsonify(goods_list)
    except Exception as e:
        logging.error(
            'An error occurred while confirming the inbound information. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})


# Route to create a new inbound record for goods.
@record_blue.route('/inbound', methods=['GET', 'POST'])
def inbound():
    try:
        if request.method == 'POST':
            goods_list = request.json
            # Get new inbound order ID
            inboundOrderID = int(Inbound.get_max_inboundOrderID()) + 1
            # Get new goods ID
            newGoodsID = int(Goods.get_max_goodsID()) + 1
            for data in goods_list:
                # Determine if the goods exist
                goodsExist = Goods.judgeExist(
                    data['goodsName'],
                    data['goodsSpecification'],
                    data['goodsManufacturer'],
                    data['goodsProductionLicense'],
                )
                #  If the goods exist, update the quantity of goods and create an inbound record
                if goodsExist != '0':
                    amount = int(
                        Goods.get_goods_by_id(goodsExist)['goodsAmount']
                    ) + int(data['goodsAmount'])
                    Goods.update_goods(
                        goodsExist,
                        data['goodsName'],
                        data['goodsSpecification'],
                        data['goodsManufacturer'],
                        data['goodsProductionLicense'],
                        data['goodsUnit'],
                        amount,
                        data['goodsStorageCondition'],
                        data['goodsUpdatedByID'],
                        datetime.now(),
                    )
                    Inbound.add_inbound(
                        str(int(Inbound.get_max_inboundID()) + 1),
                        str(inboundOrderID),
                        goodsExist,
                        data['goodsAmount'],
                        data['goodsUpdatedByID'],
                        datetime.now(),
                    )
                else:
                    # If the goods do not exist, create a goods record and an entry record
                    Goods.add_goods(
                        newGoodsID,
                        data['goodsName'],
                        data['goodsSpecification'],
                        data['goodsManufacturer'],
                        data['goodsProductionLicense'],
                        data['goodsUnit'],
                        data['goodsAmount'],
                        data['goodsStorageCondition'],
                        data['goodsUpdatedByID'],
                        datetime.now(),
                    )
                    Inbound.add_inbound(
                        str(int(Inbound.get_max_inboundID()) + 1),
                        str(inboundOrderID),
                        newGoodsID,
                        data['goodsAmount'],
                        data['goodsUpdatedByID'],
                        datetime.now(),
                    )
                    newGoodsID += 1
            return jsonify({'status': 'GET'})
    except Exception as e:
        logging.error(
            'An error occurred while creating the inbound record. Error message: {}'.format(
                str(e)
            )
        )
        return jsonify({"error": str(e)})
